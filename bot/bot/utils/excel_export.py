import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_appointments_excel(appointments_data, filename=None):
    """
    Создает Excel файл с данными записей
    Возвращает путь к созданному файлу
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"appointments_{timestamp}.xlsx"
    
    # Создаем временный файл
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)
    
    # Создаем workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Записи"
    
    # Заголовки столбцов
    headers = [
        "№", "Статус", "Дата", "Время", "Клиент", 
        "Телефон", "Услуга", "Мастер", "Цена", "Длительность"
    ]
    
    # Стиль для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Записываем данные
    for row_idx, appointment in enumerate(appointments_data, 2):
        # Парсим дату и время
        start_datetime = appointment['slot']['start']
        if 'T' in start_datetime:
            date_part = start_datetime.split('T')[0]
            time_part = start_datetime.split('T')[1][:5]
        else:
            date_part = start_datetime[:10] if len(start_datetime) >= 10 else start_datetime
            time_part = start_datetime[11:16] if len(start_datetime) >= 16 else ""
        
        # Статус
        status = "Подтверждено" if appointment.get('confirmed') else "Не подтверждено"
        
        # Данные строки
        row_data = [
            row_idx - 1,  # Номер
            status,
            date_part,
            time_part,
            appointment.get('name', ''),
            appointment.get('phone', ''),
            appointment['offering']['service']['name'],
            appointment['offering']['master']['name'],
            appointment['offering']['price'],
            format_duration_excel(appointment['offering']['duration'])
        ]
        
        # Записываем данные в строку
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(file_path)
    return file_path


def create_customers_excel(customers_data, filename=None):
    """
    Создает Excel файл с данными клиентов
    Возвращает путь к созданному файлу
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"customers_{timestamp}.xlsx"
    
    # Создаем временный файл
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)
    
    # Создаем workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    # Заголовки столбцов
    headers = [
        "№", "Статус", "Имя", "Телефон", "Дата регистрации"
    ]
    
    # Стиль для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Сортируем клиентов по статусу и дате
    sorted_customers = sorted(
        customers_data, 
        key=lambda x: (x.get('status') != 'active', x.get('created_at', '')), 
        reverse=True
    )
    
    # Записываем данные
    for row_idx, customer in enumerate(sorted_customers, 2):
        # Статус
        status = "Активный" if customer.get('status') == 'active' else "Заблокированный"
        
        # Дата регистрации
        created_date = customer.get('created_at', '')
        if 'T' in created_date:
            created_date = created_date.split('T')[0]
        
        # Данные строки
        row_data = [
            row_idx - 1,  # Номер
            status,
            customer.get('name', 'Без имени'),
            customer.get('phone', 'Без телефона'),
            created_date
        ]
        
        # Записываем данные в строку
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    wb.save(file_path)
    return file_path


def format_duration_excel(duration_str):
    """Форматирует длительность для Excel"""
    if ':' in duration_str:
        parts = duration_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1])
        
        if hours > 0:
            return f"{hours}ч {minutes}мин"
        else:
            return f"{minutes}мин"
    return duration_str


def cleanup_temp_file(file_path):
    """Удаляет временный файл"""
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"[DEBUG] Временный файл удален: {file_path}")
    except Exception as e:
        print(f"[WARN] Не удалось удалить временный файл {file_path}: {e}")